from bs4 import BeautifulSoup
import os
import requests
import socket
import json
import argparse
import sys
import os
import time
import random
from progress.bar import Bar, ChargingBar
from subprocess import *
from requests import api
import scapy.all as scapy
import nmap
import openpyxl
#from pyhunter import PyHunter
#from openpyxl import Workbook
import getpass
import smtplib
import ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart


# Funcion para la barra de progreso
def barra():
    bar2 = ChargingBar('Obteniendo Datos:', max=100)
    for num in range(100):
        time.sleep(random.uniform(0, 0.1))
        bar2.next()
    bar2.finish()


# Funcion para hacer la llamada a la API de Geolocalizacion
def geolocalizacion():
    parser = argparse.ArgumentParser()
    parser.add_argument("-ip", dest="ip", default="8.8.4.4",
                        help="Ingresa la ip, si no pones nada se tomará automatico 8.8.4.4")
    parser.add_argument("-k", dest="ip_local", help="Ingresa la ip")
    parser.add_argument("-d", dest="domain",
                        help="Ingresa un dominio para hunter")
    ip_public = parser.parse_args()
    ip_local = parser.parse_args()
    domain = parser.parse_args()
    url = "http://free.ipwhois.io/json/{}".format(ip_public.ip)
    barra()
    soup = requests.get(url)
    data = soup.text
    data = soup.json()
    datos = open("datos.txt", "w")
    #variables para excel
    n=1
	q=0
	r=0
	valor = []
	valor2= []
	wb = openpyxl.Workbook()
	ws = wb.active

    
    for key in data:
        #para guardar el txt
        datos.write(key + ": " + str(data[key])+"\n")
        #para guardar el excel
		valor.append(str(key))
		ws.cell(row=r+1,column=1).value = str(valor[q])

		valor2.append(str(data[key]))
		ws.cell(row=r+1,column=2).value = str(valor2[q])

		
		
		n=n+1
		q=q+1
		r=r+1
		
	wb.save('datoss.xlsx')
    datos.close()
    print("Archivo con datos de Geolocalizacion generado con exito ")
    return ip_local, domain


# Funcion para scanear las ip activas en una red local


def scan(ip):
    # Usamos argparser para pasar los argumentos por terminal
    archivo = open("ips.txt", "w")
    print("Scanning...")
    # iniciamos el scaneo
    arp_request = scapy.ARP(pdst=ip.ip_local+"/24")
    brodcast = scapy.Ether(dst="ff:ff:ff:ff:ff:ff")
    arp = brodcast/arp_request
    barra()
    answered = scapy.srp(arp, timeout=1, verbose=False)[0]
    # Guardamos en un txt las ip y las mac addres con un for recorriendo answered
    for element in answered:
        archivo.write("IP:{}".format(element[1].psrc))
        archivo.write("MAC address: {}\n".format(element[1].hwsrc))
    archivo.close()
    print("Archivo con ip´s activas en la red local ingresada generado con exito ")


# Se Empieza a leer el archivo que contiene las ip para almacenar en una lista


def Ports():
    lista_ip = []
    archivo = open("ips.txt", "r")
    lineas = archivo.readlines()
    for i in lineas:
        lista_ip.append((i[3:18]))
    print("Estas son las ip que puedes utilizar: -> ", lista_ip)
    archivo.close()

    # Declaramos listas
    lista_puertos = [80, 8080, 22, 23, 21, 443, 3306, 53, ]
    open_ports = []
    ip = input("ingresa la ip para scanear puertos: ")
    # Solicitamos la ip para escaenadr puertos
    #ip_add_entered = input("\nPlease enter the ip address that you want to scan: ")
    for port in lista_puertos:
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(0.5)
                s.connect((ip, port))
                open_ports.append(port)
        except:
            pass
    for port in open_ports:
        print("Port {} is open on {}".format(port, ip))
    op = input("Deseas Escanear otra ip: [y]o[n]")
    while op == "y":
        Ports()
    else:
        print("Scrip Fuera")
        exit()

# Echale un ojo a esta funcion max y carlos


def my_ip():
    url1 = 'https://www.cual-es-mi-ip.net/'
    # Peticiones a cada uno de los links
    page1 = requests.get(url1)
    soup1 = BeautifulSoup(page1.content, "html.parser")
    origen = soup1.find_all("span", class_="big-text font-arial")
    for i in origen:
        ip_public = i.text
    code = page1.status_code
    if code == 200:
        return ip_public
    else:
        ip = ''

def send_correo():
    # Lugar de trabajo de carlos
    # Correo que estamos usando paymentsnoreplybbva@gmail.com
    # Crear el mensaje
    #pedir datos de inicio de sesion
    usuario= input("Ingrese su nombre de usuario:")
    contraseña=getpass.getpass("Ingrese su  contraseña:")

    destinatario = input("Ingrese el correo del destinatario: ")
    asunto = input("Ingrese el asunto a tratar:")
    # Crear el mensaje
    cuerpoDelMensaje = MIMEMultipart("alternative")
    cuerpoDelMensaje["Subject"] = asunto
    cuerpoDelMensaje["From"] = usuario
    cuerpoDelMensaje["To"] = destinatario

    html = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta http-equiv="X-UA-Compatible" content="IE=edge">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Document</title>
        </head>
        <body>
        <h1>Hola</h2>
        </body>
        </html>

        """

    parte_html = MIMEText(html, "html")

    cuerpoDelMensaje.attach(parte_html)

    #Enviar el mensaje
    archivo = "C:/Users/Laboratorio/OneDrive - Universidad Autonoma de Nuevo León/2020-2021/Desktop/CiberTools/29-09-2021/MicrosoftTeams-image_(1).png"

    with open(archivo, "rb") as adjunto:
            contenido_adjunto = MIMEBase("aplication", "octet-stream")
            contenido_adjunto.set_payload(adjunto.read())

    encoders.encode_base64(contenido_adjunto)

    contenido_adjunto.add_header(
            "Content-Dispotition",
            f"attachment; filename = {archivo}",
    )

    cuerpoDelMensaje.attach(contenido_adjunto)
    mensaje_final = cuerpoDelMensaje.as_string()

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        server.login(usuario, contraseña)
        print("Seion Iniciada!")
        server.sendmail(usuario, destinatario, mensaje_final)
        print("Mensaje Enviado!")



    # Lugar de trabajo de Max
    #def enviarcorreo():
        # pedir datos para iniciar sesion
        #correo = input("Ingrese su correo: ")
        #contraseña = input("Ingrese su  contraseña: ")
        #destinatario = input("Ingrese el correo del destinatario: ")
        #asunto = input("Ingrese el asunto a tratar: ")
        #cuerpoDelMensaje = input("Ingresa el texto que quieras añadirle: ")
    #
        ## Crear la conexion
        #context = ssl.create_default_context()
    #
        #with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        #    server.login(correo, contraseña)
        #    print("Inicia Sesion!")
        #    #destinatario = input("Ingrese el correo del destinatario: ")
        #    #asunto = input("Ingrese el asunto a tratar :")
        #    server.sendmail(correo, destinatario, cuerpoDelMensaje, asunto)
        #    print("Mensaje enviado!")
        #    #Crear el mensaje
        #    cuerpoDelMensaje = MIMEMultipart("alternative")
        #    cuerpoDelMensaje["Subject"] = asunto
        #    cuerpoDelMensaje["From"] = correo
        #    cuerpoDelMensaje["To"] = destinatario
    #
        #html = f"""
        ## ## Aqui puedes solocar tu mensaje en HTML 
        ## 
        ##"""
    #
        #parte_html = MIMEText(html, "html")
    #
        #cuerpoDelMensaje.attach(parte_html)
    #
        #archivo = "MicrosoftTeams-image(1).png"
        #with open(archivo, "rb") as adjunto:
        #    contenido_adjunto = MIMEBase("aplication", "octet-stream")
        #    contenido_adjunto.set_payload(adjunto.read())
    #
        #encoders.encode_base64(contenido_adjunto)
    #
        #contenido_adjunto.add_header(
        #    "Content-Dispotition",
        #    f"attachment; filename = {archivo}",
        #)
    #
        #cuerpoDelMensaje.attach(contenido_adjunto)
        #mensaje_final = cuerpoDelMensaje.as_string()
    #
        #context = ssl.create_default_context()
        #with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
        #   server.login(correo, contraseña)
        #   print("Seion Iniciada!")
        #   server.sendmail(correo, destinatario, mensaje_final)
        #   print("Mensaje Enviado!")
        #return cuerpoDelMensaje
    #enviarcorreo()

    # Lugar de trabajo de Mendiola


    # Lugar de trabajo de Emilio
send_correo()